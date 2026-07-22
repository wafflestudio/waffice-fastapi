import csv
import io
from datetime import date

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models import User
from app.services import MemberService, UserService

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_CT = "text/csv"
HEADERS = ("이름", "이메일", "학번", "역할", "포지션")


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _xlsx(rows, headers=HEADERS) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _csv(rows, headers=HEADERS) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def _create_project(client, token, name, members) -> int:
    response = client.post(
        "/projects",
        json={
            "name": name,
            "started_at": date.today().isoformat(),
            "members": members,
        },
        headers=_auth(token),
    )
    assert response.status_code == 200
    return response.json()["data"]["id"]


def _replace(
    client, token, project_id, rows, filename="members.xlsx", content_type=XLSX_CT
):
    content = _csv(rows) if filename.endswith(".csv") else _xlsx(rows)
    return client.put(
        f"/projects/{project_id}/members/bulk",
        files={"file": (filename, content, content_type)},
        headers=_auth(token),
    )


def test_project_list_uses_summary_and_id_cursor(
    client: TestClient,
    admin_token: str,
    regular_token: str,
    admin_user: User,
    regular_user: User,
):
    member_rows = [
        {"user_id": admin_user.id, "role": "leader"},
        {"user_id": regular_user.id, "role": "member"},
    ]
    project_ids = [
        _create_project(client, admin_token, f"Project {index}", member_rows)
        for index in range(3)
    ]

    first = client.get("/projects?limit=2", headers=_auth(regular_token)).json()["data"]
    assert [item["id"] for item in first["items"]] == project_ids[::-1][:2]
    assert first["next_cursor"] == project_ids[1]
    assert first["items"][0] == {
        "id": project_ids[2],
        "name": "Project 2",
        "leader_names": [admin_user.name],
        "member_count": 2,
        "active_member_names": sorted([admin_user.name, regular_user.name]),
        "status": "active",
    }

    second = client.get(
        f"/projects?limit=2&cursor={first['next_cursor']}",
        headers=_auth(regular_token),
    ).json()["data"]
    assert [item["id"] for item in second["items"]] == [project_ids[0]]
    assert second["next_cursor"] is None


def test_bulk_replace_requires_admin(
    client: TestClient,
    admin_token: str,
    regular_token: str,
    regular_user: User,
):
    project_id = _create_project(
        client,
        admin_token,
        "Admin Only Bulk",
        [{"user_id": regular_user.id, "role": "leader"}],
    )

    response = _replace(
        client,
        regular_token,
        project_id,
        [(regular_user.name, regular_user.email, "", "팀장", "")],
    )
    assert response.status_code == 403


def test_template_and_bulk_replace_members_atomically(
    client: TestClient,
    db: Session,
    admin_token: str,
    admin_user: User,
    regular_user: User,
    active_user: User,
):
    active_user.student_id = "2026-0001"
    db.commit()
    project_id = _create_project(
        client,
        admin_token,
        "Roster Project",
        [
            {"user_id": admin_user.id, "role": "leader", "position": "PM"},
            {"user_id": regular_user.id, "role": "member", "position": "Backend"},
        ],
    )

    template = client.get(
        f"/projects/{project_id}/members/template", headers=_auth(admin_token)
    )
    assert template.status_code == 200
    assert template.headers["content-type"].startswith(XLSX_CT)
    workbook = load_workbook(io.BytesIO(template.content), read_only=True)
    assert list(workbook.active.values) == [
        HEADERS,
        (admin_user.name, admin_user.email, None, "팀장", "PM"),
        (regular_user.name, regular_user.email, None, "팀원", "Backend"),
    ]
    workbook.close()

    response = _replace(
        client,
        admin_token,
        project_id,
        [
            (admin_user.name, admin_user.email, "", "팀장", ""),
            (active_user.name, "", active_user.student_id, "팀원", "Frontend"),
        ],
    )
    assert response.status_code == 200
    members = response.json()["data"]["members"]
    assert {member["user"]["id"] for member in members} == {
        admin_user.id,
        active_user.id,
    }
    assert (
        next(member for member in members if member["user"]["id"] == admin_user.id)[
            "position"
        ]
        is None
    )
    assert MemberService.get_active(db, project_id, regular_user.id) is None
    assert (
        MemberService.get_active(db, project_id, active_user.id).position == "Frontend"
    )

    temporary = UserService.create(
        db,
        name="Temporary User",
        student_id="2026-temp",
        is_temporary=True,
    )
    csv_response = _replace(
        client,
        admin_token,
        project_id,
        [
            (admin_user.name, admin_user.email, "", "팀장", ""),
            (temporary.name, "", temporary.student_id, "팀원", "Designer"),
        ],
        filename="members.csv",
        content_type=CSV_CT,
    )
    assert csv_response.status_code == 200
    assert {
        member["user"]["id"] for member in csv_response.json()["data"]["members"]
    } == {
        admin_user.id,
        temporary.id,
    }


def test_invalid_bulk_file_rolls_back_every_change(
    client: TestClient,
    db: Session,
    admin_token: str,
    admin_user: User,
    regular_user: User,
):
    project_id = _create_project(
        client,
        admin_token,
        "Atomic Project",
        [
            {"user_id": admin_user.id, "role": "leader"},
            {"user_id": regular_user.id, "role": "member", "position": "Backend"},
        ],
    )
    response = _replace(
        client,
        admin_token,
        project_id,
        [
            (admin_user.name, admin_user.email, "", "팀장", ""),
            (regular_user.name, regular_user.email, "", "관리자", "Frontend"),
        ],
    )

    assert response.status_code == 400
    assert response.json()["error"] == "INVALID_PROJECT_MEMBER_FILE"
    assert response.json()["data"]["errors"][0]["code"] == "invalid_role"
    active = MemberService.list_active(db, project_id)
    assert {member.user_id for member in active} == {admin_user.id, regular_user.id}
    assert (
        MemberService.get_active(db, project_id, regular_user.id).position == "Backend"
    )


def test_bulk_rejects_identifier_mismatch_and_missing_leader(
    client: TestClient,
    db: Session,
    admin_token: str,
    admin_user: User,
    active_user: User,
):
    active_user.student_id = "2026-9999"
    db.commit()
    project_id = _create_project(
        client,
        admin_token,
        "Validation Project",
        [{"user_id": admin_user.id, "role": "leader"}],
    )

    mismatch = _replace(
        client,
        admin_token,
        project_id,
        [(admin_user.name, admin_user.email, active_user.student_id, "팀장", "")],
    )
    assert mismatch.status_code == 400
    assert mismatch.json()["data"]["errors"][0]["code"] == "identifier_mismatch"

    no_leader = _replace(
        client,
        admin_token,
        project_id,
        [(admin_user.name, admin_user.email, "", "팀원", "")],
    )
    assert no_leader.status_code == 400
    assert no_leader.json()["data"]["errors"][0]["code"] == "no_leader"
    assert (
        MemberService.get_active(db, project_id, admin_user.id).role.value == "leader"
    )


def test_member_patch_distinguishes_omitted_and_null_position(
    client: TestClient,
    admin_token: str,
    admin_user: User,
    regular_user: User,
):
    project_id = _create_project(
        client,
        admin_token,
        "Patch Project",
        [
            {"user_id": admin_user.id, "role": "leader"},
            {"user_id": regular_user.id, "role": "member", "position": "Backend"},
        ],
    )

    role_only = client.patch(
        f"/projects/{project_id}/members/{regular_user.id}",
        json={"role": "leader"},
        headers=_auth(admin_token),
    )
    assert role_only.status_code == 200
    member = next(
        item
        for item in role_only.json()["data"]["members"]
        if item["user"]["id"] == regular_user.id
    )
    assert member["position"] == "Backend"

    clear_position = client.patch(
        f"/projects/{project_id}/members/{regular_user.id}",
        json={"position": None},
        headers=_auth(admin_token),
    )
    assert clear_position.status_code == 200
    member = next(
        item
        for item in clear_position.json()["data"]["members"]
        if item["user"]["id"] == regular_user.id
    )
    assert member["position"] is None
