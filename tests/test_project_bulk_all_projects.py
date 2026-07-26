import io
from datetime import date

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models import ActivityStatus, ProjectMember, User, UserActivity
from app.services import MemberService

XLSX_CT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HEADERS = ("프로젝트명", "프로젝트원 이름", "학번", "팀장 여부", "포지션")


def _auth(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _xlsx(rows, headers=HEADERS) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _create_project(client, token, name, members) -> int:
    response = client.post(
        "/projects",
        json={
            "name": name,
            "started_at": date.today().isoformat(),
            "members": members,
        },
        headers=_auth(token),
    )
    assert response.status_code == 200
    return response.json()["data"]["id"]


def _upload(client, token, rows, filename="members.xlsx"):
    content = _xlsx(rows)
    return client.put(
        "/projects/members/bulk-all-projects",
        files={"file": (filename, content, XLSX_CT)},
        headers=_auth(token),
    )


def test_requires_admin(client: TestClient, regular_token: str, admin_user: User):
    response = _upload(
        client, regular_token, [("Anything", admin_user.name, "0000", "O", "")]
    )
    assert response.status_code == 403


def test_rejects_non_xlsx_file(client: TestClient, admin_token: str):
    response = client.put(
        "/projects/members/bulk-all-projects",
        files={"file": ("members.csv", b"whatever", "text/csv")},
        headers=_auth(admin_token),
    )
    assert response.status_code == 400
    assert (
        response.json()["data"]["errors"][0]["message"] == ".xlsx 파일을 첨부해주세요."
    )


def test_missing_header_reports_column_name(client: TestClient, admin_token: str):
    # Drop the 학번 header to trigger a missing_header error.
    content = _xlsx(
        [("Project", "Name", "O", "")],
        headers=("프로젝트명", "프로젝트원 이름", "팀장 여부", "포지션"),
    )
    response = client.put(
        "/projects/members/bulk-all-projects",
        files={"file": ("members.xlsx", content, XLSX_CT)},
        headers=_auth(admin_token),
    )
    assert response.status_code == 400
    errors = response.json()["data"]["errors"]
    assert any(
        error["code"] == "missing_header"
        and error["message"] == "학번 열을 찾을 수 없습니다."
        for error in errors
    )


def test_project_not_found(
    client: TestClient, admin_token: str, admin_user: User, db: Session
):
    admin_user.student_id = "2030-0001"
    db.commit()
    response = _upload(
        client,
        admin_token,
        [("존재하지않는프로젝트", admin_user.name, admin_user.student_id, "O", "")],
    )
    assert response.status_code == 400
    error = response.json()["data"]["errors"][0]
    assert error["code"] == "project_not_found"
    assert (
        error["message"]
        == "존재하지않는프로젝트 프로젝트명을 찾을 수 없습니다. 프로젝트 목록에 존재하는지 확인해주세요."
    )


def test_student_id_must_belong_to_active_qualification(
    client: TestClient,
    admin_token: str,
    admin_user: User,
    regular_user: User,
    db: Session,
):
    project_id = _create_project(
        client,
        admin_token,
        "Qual Project",
        [{"user_id": admin_user.id, "role": "leader"}],
    )
    regular_user.student_id = "2030-1111"
    db.commit()

    response = _upload(
        client,
        admin_token,
        [
            (
                "Qual Project",
                admin_user.name,
                admin_user.student_id or "",
                "O",
                "",
            ),
            ("Qual Project", regular_user.name, regular_user.student_id, "X", ""),
        ],
    )
    assert response.status_code == 400
    errors = response.json()["data"]["errors"]
    assert any(
        error["code"] == "user_not_found"
        and error["message"]
        == f"{regular_user.name}({regular_user.student_id})을 활동회원 명부에서 찾을 수 없습니다."
        for error in errors
    )
    active_ids = {
        member.user_id for member in MemberService.list_active(db, project_id)
    }
    assert active_ids == {admin_user.id}


def test_name_mismatch(
    client: TestClient, admin_token: str, admin_user: User, db: Session
):
    admin_user.student_id = "2030-2222"
    db.commit()
    project_id = _create_project(
        client,
        admin_token,
        "Mismatch Project",
        [{"user_id": admin_user.id, "role": "leader"}],
    )
    response = _upload(
        client,
        admin_token,
        [("Mismatch Project", "다른이름", admin_user.student_id, "O", "")],
    )
    assert response.status_code == 400
    error = response.json()["data"]["errors"][0]
    assert error["code"] == "name_mismatch"
    assert error["message"] == (
        f"파일의 데이터(다른이름, {admin_user.student_id})가 DB의 데이터"
        f"({admin_user.name}, {admin_user.student_id})와 일치하지 않습니다. 오타를 확인해주세요."
    )


def test_ambiguous_student_id(
    client: TestClient,
    admin_token: str,
    admin_user: User,
    active_user: User,
    db: Session,
):
    admin_user.student_id = "2030-3333"
    active_user.student_id = "2030-3333"
    db.commit()
    project_id = _create_project(
        client,
        admin_token,
        "Ambiguous Project",
        [{"user_id": admin_user.id, "role": "leader"}],
    )
    response = _upload(
        client,
        admin_token,
        [("Ambiguous Project", admin_user.name, "2030-3333", "O", "")],
    )
    assert response.status_code == 400
    assert response.json()["data"]["errors"][0]["code"] == "ambiguous_student_id"


def test_duplicate_user_within_same_project(
    client: TestClient, admin_token: str, admin_user: User, db: Session
):
    admin_user.student_id = "2030-4444"
    db.commit()
    _create_project(
        client,
        admin_token,
        "Dup Project",
        [{"user_id": admin_user.id, "role": "leader"}],
    )
    response = _upload(
        client,
        admin_token,
        [
            ("Dup Project", admin_user.name, admin_user.student_id, "O", ""),
            ("Dup Project", admin_user.name, admin_user.student_id, "X", "Extra"),
        ],
    )
    assert response.status_code == 400
    assert response.json()["data"]["errors"][0]["code"] == "duplicate_user"


def test_no_leader_in_group(
    client: TestClient, admin_token: str, admin_user: User, db: Session
):
    admin_user.student_id = "2030-5555"
    db.commit()
    _create_project(
        client,
        admin_token,
        "No Leader Project",
        [{"user_id": admin_user.id, "role": "leader"}],
    )
    response = _upload(
        client,
        admin_token,
        [("No Leader Project", admin_user.name, admin_user.student_id, "X", "")],
    )
    assert response.status_code == 400
    assert response.json()["data"]["errors"][0]["code"] == "no_leader"


def test_diff_logic_creates_updates_and_closes_activities(
    client: TestClient,
    db: Session,
    admin_token: str,
    admin_user: User,
    active_user: User,
    regular_user: User,
):
    admin_user.student_id = "2030-6001"
    active_user.student_id = "2030-6002"
    db.commit()

    project_id = _create_project(
        client,
        admin_token,
        "Diff Project",
        [
            {"user_id": admin_user.id, "role": "leader", "position": "PM"},
            {"user_id": regular_user.id, "role": "member", "position": "Old"},
        ],
    )
    other_project_id = _create_project(
        client,
        admin_token,
        "Untouched Project",
        [{"user_id": admin_user.id, "role": "leader"}],
    )

    admin_before = MemberService.get_active(db, project_id, admin_user.id)
    admin_joined_at_before = admin_before.joined_at

    # New roster for "Diff Project": admin stays (unchanged), regular_user is
    # dropped, active_user is newly added.
    response = _upload(
        client,
        admin_token,
        [
            ("Diff Project", admin_user.name, admin_user.student_id, "O", "PM"),
            (
                "Diff Project",
                active_user.name,
                active_user.student_id,
                "X",
                "Backend",
            ),
        ],
    )
    assert response.status_code == 200
    projects = response.json()["data"]
    assert {p["id"] for p in projects} == {project_id}

    # admin_user: unchanged membership period, still active.
    admin_after = MemberService.get_active(db, project_id, admin_user.id)
    assert admin_after is not None
    assert admin_after.joined_at == admin_joined_at_before

    # regular_user: removed (left_at set), and their UserActivity closed.
    assert MemberService.get_active(db, project_id, regular_user.id) is None
    removed_member = (
        db.query(ProjectMember)
        .filter(
            ProjectMember.project_id == project_id,
            ProjectMember.user_id == regular_user.id,
        )
        .first()
    )
    assert removed_member.left_at is not None

    # active_user: newly added, with a matching open UserActivity.
    new_member = MemberService.get_active(db, project_id, active_user.id)
    assert new_member is not None
    assert new_member.position == "Backend"
    activity = (
        db.query(UserActivity)
        .filter(
            UserActivity.user_id == active_user.id,
            UserActivity.project_id == project_id,
        )
        .first()
    )
    assert activity is not None
    assert activity.position == "Backend"
    assert activity.end_date is None
    assert activity.status == ActivityStatus.ACTIVE

    # The other project, not mentioned in the file, is untouched.
    assert MemberService.get_active(db, other_project_id, admin_user.id) is not None


def test_atomic_rollback_across_projects(
    client: TestClient,
    db: Session,
    admin_token: str,
    admin_user: User,
    active_user: User,
):
    admin_user.student_id = "2030-8001"
    active_user.student_id = "2030-8002"
    db.commit()

    project_a = _create_project(
        client, admin_token, "Project A", [{"user_id": admin_user.id, "role": "leader"}]
    )
    project_b = _create_project(
        client,
        admin_token,
        "Project B",
        [{"user_id": active_user.id, "role": "leader"}],
    )

    response = _upload(
        client,
        admin_token,
        [
            # Valid change to Project A: add active_user as a member.
            ("Project A", admin_user.name, admin_user.student_id, "O", ""),
            ("Project A", active_user.name, active_user.student_id, "X", "Backend"),
            # Invalid: Project B ends up with no leader.
            ("Project B", active_user.name, active_user.student_id, "X", ""),
        ],
    )
    assert response.status_code == 400
    assert response.json()["data"]["errors"][0]["code"] == "no_leader"

    # Nothing changed in Project A either (atomic across the whole file).
    assert MemberService.get_active(db, project_a, active_user.id) is None
    assert MemberService.get_active(db, project_b, active_user.id) is not None


def test_download_template(
    client: TestClient, admin_token: str, admin_user: User, db: Session
):
    admin_user.student_id = "2030-9001"
    db.commit()
    _create_project(
        client,
        admin_token,
        "Template Project",
        [{"user_id": admin_user.id, "role": "leader", "position": "PM"}],
    )
    response = client.get(
        "/projects/members/bulk-all-projects/template", headers=_auth(admin_token)
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(XLSX_CT)
    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    rows = list(workbook.active.values)
    assert rows[0] == HEADERS
    assert (
        "Template Project",
        admin_user.name,
        admin_user.student_id,
        "O",
        "PM",
    ) in rows[1:]
    workbook.close()
