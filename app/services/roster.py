from __future__ import annotations

import csv
import io
import re
import zipfile
from typing import Iterator, NamedTuple, Sequence
from xml.etree.ElementTree import ParseError

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.exceptions import EmptyRosterError, InvalidRosterFileError, RosterTooLargeError
from app.models import MemberRole

MAX_ROWS = 2000
MAX_ROSTER_FILE_BYTES = 5 * 1024 * 1024
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_CONTENT_TYPE = "text/csv"
_MAX_NAME = 100
_MAX_EMAIL = 255
_MAX_STUDENT_ID = 50
_MAX_POSITION = 50

# Header aliases (normalized: lowercased, non-alphanumeric/non-Hangul stripped).
_NAME_HEADERS = {"이름", "성명", "성함", "name"}
_STUDENT_ID_HEADERS = {"학번", "studentid", "sid", "학번sid"}

# Zero-width / format chars that str.strip() does not treat as whitespace (e.g. BOM).
_ZERO_WIDTH = "\u200b\u200c\u200d\u2060\ufeff\u00ad"

PROJECT_MEMBER_HEADERS = ("이름", "이메일", "학번", "역할", "포지션")
_PROJECT_MEMBER_ROLES = {"팀장": MemberRole.LEADER, "팀원": MemberRole.MEMBER}


class ProjectMemberRosterRow(NamedTuple):
    row_number: int
    name: str
    email: str
    student_id: str
    role: MemberRole
    position: str | None


def _norm_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^0-9a-z가-힣]", "", str(value).strip().lower())


def _cell_to_str(value: object) -> str:
    """Render a cell as a trimmed string (int/float ids -> digits, not '2021.0')."""
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _normalize(value: str) -> str:
    return value.strip().strip(_ZERO_WIDTH).strip()


def _xlsx_rows(content: bytes) -> Iterator[Sequence[object]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (
        InvalidFileException,
        zipfile.BadZipFile,
        KeyError,
        OSError,
        ValueError,
        ParseError,
    ):
        raise InvalidRosterFileError()
    try:
        yield from workbook.active.iter_rows(values_only=True)
    finally:
        workbook.close()


def _csv_rows(content: bytes) -> Iterator[Sequence[object]]:
    for encoding in ("utf-8-sig", "cp949"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            text = None
    if text is None:
        raise InvalidRosterFileError()
    yield from csv.reader(io.StringIO(text))


def parse_member_roster(
    content: bytes,
    filename: str,
) -> tuple[list[tuple[str, str]], list[tuple[str, str, str]]]:
    """
    Parse an .xlsx or .csv member roster into (valid_rows, invalid_rows).

    The file type is chosen by the filename extension (.xlsx or .csv); any other
    extension is rejected. The first row is the header; the name and student_id
    columns are located by header text (case-insensitive, Korean/English aliases,
    column order agnostic):
      - name:       이름 / 성명 / name
      - student_id: 학번 / student_id / sid

    Returns:
      valid_rows:   list of (name, student_id), normalized.
      invalid_rows: list of (name, student_id, reason) for rows missing a field —
                    reason is "missing_name", "missing_student_id", or "invalid"
                    (over-length). Skipped, not fatal.

    Raises:
      InvalidRosterFileError (400): unsupported extension, unreadable file, or
                                    missing name/student_id header.
      EmptyRosterError (422): header present but no data rows.
      RosterTooLargeError (400): more than MAX_ROWS data rows.
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        rows = _csv_rows(content)
    elif name.endswith(".xlsx"):
        rows = _xlsx_rows(content)
    else:
        raise InvalidRosterFileError()

    try:
        header = next(rows)
    except StopIteration:
        raise EmptyRosterError()

    name_idx = student_id_idx = None
    for index, cell in enumerate(header or ()):
        key = _norm_header(cell)
        if name_idx is None and key in _NAME_HEADERS:
            name_idx = index
        if student_id_idx is None and key in _STUDENT_ID_HEADERS:
            student_id_idx = index

    if name_idx is None:
        raise InvalidRosterFileError("이름 헤더를 찾을 수 없습니다.")
    if student_id_idx is None:
        raise InvalidRosterFileError("학번 헤더를 찾을 수 없습니다.")

    valid: list[tuple[str, str]] = []
    invalid: list[tuple[str, str, str]] = []
    count = 0
    for row in rows:
        name = _normalize(_cell_to_str(row[name_idx]) if name_idx < len(row) else "")
        student_id = _normalize(
            _cell_to_str(row[student_id_idx]) if student_id_idx < len(row) else ""
        )
        if not name and not student_id:
            continue  # fully-empty spacer row
        count += 1
        if count > MAX_ROWS:
            raise RosterTooLargeError()

        if not student_id:
            invalid.append((name, student_id, "missing_student_id"))
        elif not name:
            invalid.append((name, student_id, "missing_name"))
        elif len(name) > _MAX_NAME or len(student_id) > _MAX_STUDENT_ID:
            invalid.append((name, student_id, "invalid"))
        else:
            valid.append((name, student_id))

    if not valid and not invalid:
        raise EmptyRosterError()

    return valid, invalid


def parse_project_member_roster(
    content: bytes, filename: str
) -> tuple[list[ProjectMemberRosterRow], list[dict]]:
    """Parse the fixed Korean project-member XLSX/CSV format."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        rows = _csv_rows(content)
    elif name.endswith(".xlsx"):
        rows = _xlsx_rows(content)
    else:
        return [], [
            _project_member_error(
                0,
                "file",
                "invalid_file",
                ".xlsx 또는 .csv 파일을 첨부해주세요.",
            )
        ]

    try:
        header = next(rows)
    except (InvalidRosterFileError, StopIteration):
        return [], [
            _project_member_error(0, "file", "invalid_file", "파일 양식이 올바르지 않습니다.")
        ]

    indexes = {_norm_header(cell): index for index, cell in enumerate(header or ())}
    errors = [
        _project_member_error(
            1, header_name, "missing_header", f"{header_name} 열을 찾을 수 없습니다."
        )
        for header_name in PROJECT_MEMBER_HEADERS
        if _norm_header(header_name) not in indexes
    ]
    if errors:
        return [], errors

    parsed: list[ProjectMemberRosterRow] = []
    count = 0
    for row_number, row in enumerate(rows, start=2):
        values = {
            header_name: _normalize(
                _cell_to_str(row[indexes[_norm_header(header_name)]])
                if indexes[_norm_header(header_name)] < len(row)
                else ""
            )
            for header_name in PROJECT_MEMBER_HEADERS
        }
        if not any(values.values()):
            continue

        count += 1
        if count > MAX_ROWS:
            errors.append(
                _project_member_error(
                    row_number,
                    "file",
                    "too_many_rows",
                    f"명단은 최대 {MAX_ROWS}행까지 가능합니다.",
                )
            )
            break

        row_errors: list[dict] = []
        if not values["이름"]:
            row_errors.append(
                _project_member_error(row_number, "이름", "required", "이름을 입력해주세요.")
            )
        elif len(values["이름"]) > _MAX_NAME:
            row_errors.append(
                _project_member_error(
                    row_number,
                    "이름",
                    "too_long",
                    f"이름은 {_MAX_NAME}자 이하여야 합니다.",
                )
            )

        if not values["이메일"] and not values["학번"]:
            row_errors.append(
                _project_member_error(
                    row_number,
                    "이메일",
                    "missing_identifier",
                    "이메일 또는 학번을 입력해주세요.",
                )
            )
        if len(values["이메일"]) > _MAX_EMAIL:
            row_errors.append(
                _project_member_error(
                    row_number,
                    "이메일",
                    "too_long",
                    f"이메일은 {_MAX_EMAIL}자 이하여야 합니다.",
                )
            )
        if len(values["학번"]) > _MAX_STUDENT_ID:
            row_errors.append(
                _project_member_error(
                    row_number,
                    "학번",
                    "too_long",
                    f"학번은 {_MAX_STUDENT_ID}자 이하여야 합니다.",
                )
            )

        role = _PROJECT_MEMBER_ROLES.get(values["역할"])
        if role is None:
            row_errors.append(
                _project_member_error(
                    row_number,
                    "역할",
                    "invalid_role",
                    "역할은 팀장 또는 팀원이어야 합니다.",
                )
            )
        if len(values["포지션"]) > _MAX_POSITION:
            row_errors.append(
                _project_member_error(
                    row_number,
                    "포지션",
                    "too_long",
                    f"포지션은 {_MAX_POSITION}자 이하여야 합니다.",
                )
            )

        if row_errors:
            errors.extend(row_errors)
            continue
        parsed.append(
            ProjectMemberRosterRow(
                row_number,
                values["이름"],
                values["이메일"].lower(),
                values["학번"],
                role,
                values["포지션"] or None,
            )
        )

    if not parsed and not errors:
        errors.append(
            _project_member_error(2, "file", "empty_roster", "팀원 명단이 비어 있습니다.")
        )
    return parsed, errors


def build_project_member_template(members: Sequence[object]) -> bytes:
    """Build the editable workbook from active memberships."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "팀원"
    sheet.append(PROJECT_MEMBER_HEADERS)
    for member in sorted(
        members,
        key=lambda item: (
            item.role != MemberRole.LEADER,
            item.user.name,
            item.user_id,
        ),
    ):
        sheet.append(
            (
                member.user.name,
                member.user.email or "",
                member.user.student_id or "",
                "팀장" if member.role == MemberRole.LEADER else "팀원",
                member.position or "",
            )
        )
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _project_member_error(row: int, field: str, code: str, message: str) -> dict:
    return {"row": row, "field": field, "code": code, "message": message}
